import os
import argparse
import logging
from logdecorator import log_on_start, log_on_end, log_on_error
import dns.message
import dns.rdataclass
import dns.rdatatype
import dns.query
import socket
from ping3 import ping
from datetime import datetime, timezone
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import styles
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

import ipaddress

# urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

argp = argparse.ArgumentParser()
argp.add_argument("-f", dest="dns_file", help="DNS Excel file to check, default 'DNS_Entries.xlsx'",
                  default="DNS_Entries.xlsx")
argp.add_argument("--d", default=False, action='store_true',  help="Enable Debug")
argp.add_argument("--env", default="prod",  help="Environment")

args = argp.parse_args()
if args.d:
    logging.basicConfig(level=logging.DEBUG)
    logging.basicConfig(level=logging.INFO)
else:
    logging.basicConfig(level=logging.INFO)

dns_list_global = []
dns_conn_cache = {}


@log_on_start(logging.DEBUG, "{callable.__name__:s} - {dns_file}")
@log_on_error(logging.ERROR, "Error on loading {dns_file}: {e!r}",
              reraise=True)
# @log_on_error(logging.ERROR, "{callable.__name__:s}")
@log_on_end(logging.DEBUG, "{callable.__name__:s}")
def read_dns_records(dns_file):
    df = pd.read_excel(dns_file, sheet_name="DNS_List")
    return df


@log_on_start(logging.DEBUG, "{callable.__name__:s} - {target}")
@log_on_error(logging.ERROR, "Error on {target} {check_type}: {e!r}",
              reraise=True)
@log_on_end(logging.DEBUG, "{callable.__name__:s} - {target} - {result}")
def check_ping(target):
    ping_result = ping(target, timeout=3, unit="s")
    if ping_result is None:
        result = "TIMED OUT"
    elif not ping_result:
        result = "HOST UNKNOWN"
    else:
        result = "OK"

    return result


@log_on_start(logging.DEBUG, "{callable.__name__:s} - {target}:{port}")
@log_on_error(logging.ERROR, "Error on {target} {check_type}: {e!r}",
              reraise=True)
@log_on_end(logging.DEBUG, "{callable.__name__:s} - {target}:{port} - {result}")
def check_tcp(target, port):
    assert int(port)
    timeout_seconds = 2
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(timeout_seconds)
    result_tcp = sock.connect_ex((target, port))
    if result_tcp == 0:
        result = "OK"
    else:
        result = "ERROR-{}".format(result_tcp)
    sock.close()
    return result


@log_on_start(logging.DEBUG, "{callable.__name__:s} {target} {tcp_check} {ping_check}")
@log_on_error(logging.ERROR, "Error : {e!r}",
              reraise=True)
@log_on_end(logging.DEBUG, "{callable.__name__:s}")
def check_target_conn(target, tcp_check, ping_check, dns_cache):
    logging.debug("{} - {} - {}".format(target, tcp_check, ping_check))
    result = {"check_ping": "N/A", "check_tcp": "N/A"}
    if ping_check:
        result["check_ping"] = check_ping(target)
    if type(tcp_check) == int and 0 < tcp_check < 65536:
        result["check_tcp"] = check_tcp(target, tcp_check)
    dns_cache[target] = result
    return


@log_on_start(logging.DEBUG, "{callable.__name__:s} ")
@log_on_error(logging.ERROR, "Error : {e!r}",
              reraise=True)
@log_on_end(logging.DEBUG, "{callable.__name__:s}")
def check_connectivity(df, dns_cache):
    df.apply(lambda row: check_target_conn(row['Target'], row['TCP'], row['Ping'], dns_cache), axis=1)
    # logging.debug(df.head())


@log_on_start(logging.DEBUG, "{callable.__name__:s} - {dns_entry} - {dns_target}")
@log_on_error(logging.ERROR, "Error on {dns_entry}: {e!r}",
              reraise=True)
@log_on_end(logging.DEBUG, "{callable.__name__:s} - {dns_entry} {result}")
def validate_entry(dns_entry, dns_target):
    result = "DIFF"
    if dns_entry.casefold() == dns_target.casefold():
        result = "MATCH"

    return result


# @log_on_start(logging.DEBUG, "{callable.__name__:s} ")
# @log_on_error(logging.ERROR, "Error : {e!r}",
#               reraise=True)
# @log_on_end(logging.DEBUG, "{callable.__name__:s} ")
def proccess_dns_entry(row, target_conn, dns_list):
    logging.debug("Processing {}".format(row["DNS"]))
    ns_list = []
    try:
        ns_list.append(ipaddress.ip_address(row['NS1']))
    except ValueError as e:
        logging.error("Error for {} - NS1 bad IP Address {}. {}".format(row["DNS"], row["NS1"], e))
        pass
    try:
        ns_list.append(ipaddress.ip_address(row['NS2']))
    except ValueError as e:
        logging.error("Error for {} - NS1 bad IP Address {}. {}".format(row["DNS"], row["NS2"], e))
        pass

    for ns in ns_list:
        logging.debug("Using NS: {}".format(ns))
        result = {'Env': row['Env'], "DNS":  row['DNS'], "Type": row['Type'], "Target": row['Target'], "NS": str(ns),
                  "Status": "NO_MATCH", "TCP_Connectivity": "N/A", "Ping": "N/A", "DNS_ANSWER": ""}
        dns_type = dns.rdatatype.CNAME
        if row['Type'] == "A":
            dns_type = dns.rdatatype.A
        elif row['Type'] != "CNAME":
            logging.error("For {} the DNS Type {} is not supported".format(row['DNS'], row['Type']))
            continue

        qname = dns.name.from_text(row['DNS'])
        q = dns.message.make_query(qname, dns_type)
        try:
            r = dns.query.udp(q, str(ns), timeout=5)
            if len(r.answer) == 0:
                logging.error("NO DNS: {}".format(qname))
                result["Status"] = "NO_EXIST"
            else:
                ns_rrset = r.find_rrset(r.answer, qname, dns.rdataclass.IN, dns_type)

                for rr in ns_rrset:
                    dns_target = ""
                    if row["Type"] == "A":
                        dns_target = rr.address
                    elif row["Type"] == "CNAME":
                        dns_target = rr.target
                    logging.debug("Found {} {} {}".format(row["DNS"], row["Type"], row["Target"], dns_target))
                    dns_target = str(dns_target).rstrip(".")
                    if validate_entry(row["Target"], dns_target) == "MATCH":
                        result["Status"] = "MATCH"
                        result["TCP_Connectivity"] = target_conn["check_tcp"]
                        result["Ping"] = target_conn["check_ping"]
                    result["DNS_ANSWER"] = "{}\n".format(dns_target)

                result["Status"] = result["Status"].rstrip('\n')
            dns_list.append(result)
        # except Exception as eDnsTimeout:
        except dns.exception.Timeout as eDnsTimeout:
            logging.error("NS {} not exists or not available.".format(ns))
            pass


@log_on_start(logging.DEBUG, "{callable.__name__:s}")
@log_on_error(logging.ERROR, "Error on  {e!r}", reraise=True)
@log_on_end(logging.DEBUG, "{callable.__name__:s}")
def main_dns_process(df, target_cache, dns_list):
    df.apply(lambda row: proccess_dns_entry(row, target_cache[row['Target']], dns_list), axis=1)


dns_to_check = read_dns_records(args.dns_file)
dns_to_check_env = dns_to_check[dns_to_check['Env'] == args.env]
check_connectivity(dns_to_check_env[['Target', 'TCP', 'Ping']].drop_duplicates(), dns_conn_cache)
main_dns_process(dns_to_check_env, dns_conn_cache, dns_list_global)

logging.debug(dns_conn_cache)
logging.debug(dns_list_global)
df_result = pd.DataFrame(dns_list_global)

book = load_workbook(args.dns_file)
with pd.ExcelWriter("{}.xlsx".format(os.path.splitext(args.dns_file)[0]), engine='openpyxl') as writer:
    writer.book = book
    sheet_name = 'CHECK_{}_{}'.format(args.env, datetime.now(timezone.utc).strftime("%Y%m%d%M%S"))
    df_result.to_excel(writer, sheet_name=sheet_name, index=False)
    tab = Table(displayName="TableResult_{}".format(sheet_name), ref="A1:I{}".format(len(df_result.index) + 1))
    style = TableStyleInfo(name="TableStyleLight14", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    red_color = 'ffc7ce'
    red_color_font = '9c0103'
    cond_matrix = "F2:F{}".format(len(df_result.index) + 1)
    all_matrix = "A2:I{}".format(len(df_result.index) + 1)
    red_font = styles.Font(size=14, bold=True, color=red_color_font)
    red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')

    rule = Rule(type='expression', formula=['NOT($F2="MATCH")'], stopIfTrue=True)
    rule.dxf = DifferentialStyle(font=red_font, border=None, fill=red_fill)
    writer.book[sheet_name].conditional_formatting.add("{}".format(all_matrix), rule)

    writer.book[sheet_name].add_table(tab)

